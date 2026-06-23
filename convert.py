#!/usr/bin/env python3
"""訪問内容一覧表.xlsx をスマホ閲覧用 index.html に変換する。

使い方:
    python convert.py 訪問内容一覧表.xlsx [出力先index.html]
"""
import json
import re
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

from openpyxl import load_workbook

DAYS = ["月", "火", "水", "木", "金", "土", "日"]
CONTACT_SHEET = "連絡先"


def parse_sheet(ws, sheet_name):
    client = {"name": sheet_name.strip(), "policy": "", "visits": {d: [] for d in DAYS}, "notes": []}
    current_day = None
    for row in ws.iter_rows(min_row=1):
        b, c, d, e, f = (row[i].value if i < len(row) else None for i in range(1, 6))
        b = str(b).strip() if b else ""
        c = str(c).strip() if c else ""
        if b == "【援助方針】":
            client["policy"] = str(row[2].value or "").strip()
            continue
        if c == "曜日":  # ヘッダー行
            continue
        if b.startswith("※"):
            client["notes"].append(b.lstrip("※").strip())
            continue
        day = next((x for x in DAYS if c.startswith(x)), None)
        if day:
            current_day = day
        if current_day and d:
            client["visits"][current_day].append({
                "time": str(d).strip(),
                "code": str(e or "").strip(),
                "text": str(f or "").strip(),
            })
    return client


def parse_contacts(ws):
    """「連絡先」シートを読み、{利用者名: {type: raw_value}} を返す。"""
    data = {}
    header_names = []
    first = True
    for row in ws.iter_rows(min_row=1):
        if first:
            first = False
            # B列（index 1）以降が利用者名
            for cell in row[1:]:
                name = str(cell.value).strip() if cell.value else ""
                header_names.append(name)
            continue
        row_type = str(row[0].value).strip() if row[0].value else ""
        if row_type not in ["自宅", "家族", "担当ケアマネ"]:
            continue
        for col_idx, cell in enumerate(row[1:]):
            if col_idx >= len(header_names):
                break
            client_name = header_names[col_idx]
            if not client_name:
                continue
            val = str(cell.value).strip() if cell.value else ""
            if val and val != "（空欄）":
                if client_name not in data:
                    data[client_name] = {}
                data[client_name][row_type] = val
    return data


def make_phone_entry(row_type, val):
    """連絡先の種類と値から phones エントリを作る。"""
    if "・" in val:
        parts = val.split("・", 1)
        name = parts[0].strip()
        display = parts[1].strip()
    else:
        name = ""
        display = val.strip()
    tel = re.sub(r"\D", "", display)
    return {"type": row_type, "name": name, "display": display, "tel": tel}


def build_phones(sheet_name, contacts_data):
    """シート名から phones リストを作る（連絡先シートの列見出しと突合）。"""
    # シート名の「・」前を基本名として使う
    base_name = sheet_name.split("・")[0].strip() if "・" in sheet_name else sheet_name.strip()
    raw = contacts_data.get(base_name) or contacts_data.get(sheet_name.strip())
    if not raw:
        return []
    phones = []
    for row_type in ["自宅", "家族", "担当ケアマネ"]:
        val = raw.get(row_type, "")
        if val:
            phones.append(make_phone_entry(row_type, val))
    return phones


def main():
    if len(sys.argv) < 2:
        sys.exit("使い方: python convert.py 訪問内容一覧表.xlsx [index.html]")
    src = Path(sys.argv[1])
    dst = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(__file__).parent / "index.html"

    wb = load_workbook(src, data_only=True)

    # 連絡先シート読み込み
    contacts_data = {}
    if CONTACT_SHEET in wb.sheetnames:
        contacts_data = parse_contacts(wb[CONTACT_SHEET])
        print(f"連絡先シート読み込み: {len(contacts_data)}件")
    else:
        print(f"警告: 「{CONTACT_SHEET}」シートが見つかりません。電話ボタンは表示されません。")

    # 利用者シート読み込み（連絡先シートを除外）
    client_sheet_names = [name for name in wb.sheetnames if name != CONTACT_SHEET]
    clients = []
    unmatched = []
    for name in client_sheet_names:
        client = parse_sheet(wb[name], name)
        client["phones"] = build_phones(name, contacts_data)
        if contacts_data and not client["phones"]:
            unmatched.append(name)
        clients.append(client)

    if unmatched:
        print(f"未突合の利用者（電話ボタン非表示）: {', '.join(unmatched)}")

    jst = timezone(timedelta(hours=9))
    updated = datetime.now(jst).strftime("%Y/%m/%d %H:%M")

    template = (Path(__file__).parent / "template.html").read_text(encoding="utf-8")
    html = template.replace("/*__DATA__*/[]", json.dumps(clients, ensure_ascii=False))
    html = html.replace("__UPDATED__", updated)
    dst.write_text(html, encoding="utf-8")
    print(f"OK: {dst}（利用者{len(clients)}名、更新 {updated}）")


if __name__ == "__main__":
    main()
